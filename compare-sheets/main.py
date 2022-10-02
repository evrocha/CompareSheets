from cgi import print_form
from pdb import line_prefix
from re import X
from xml.etree.ElementTree import tostring
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, QListWidgetItem, QPushButton
from PyQt5.QtCore import Qt
from PyQt5.uic import loadUi
import sys;

import pandas as pd
from pandas import *

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import *
from IPython.display import display
sheet1 = []
sheet2 = []
wblist = []
paths = []
list_qtd = []
list_cod = []
list_qtd2 = []
list_cod2 = []
#                                                 --- PANDAS VERSION
class MainWindow(QDialog):
    
    def __init__(self):
        super(MainWindow, self).__init__()
        super().__init__()
        
        loadUi("D:\Programação\CompareSheets\compare-sheets\\up_file-screen.ui", self)
        self.browse.clicked.connect(self.get_first_sheet)
        self.browse2.clicked.connect(self.get_sec_sheet)
        self.comparar_btn.clicked.connect(self.comparar)

    def get_first_sheet(self):
        get_version_old = QFileDialog.getOpenFileName(self, 'Open file', 'D:\Programação\CompareSheets\compare-sheets')         # pega o arquivo
        self.filename.setText(get_version_old[0])  
        file_path = (get_version_old[0]) # caminho 

        paths.append(file_path)

        initial_sheet = pd.read_excel(file_path) 
        sheet1.append(initial_sheet)

        planilha01 = pd.read_excel(file_path)
        planilha01 = planilha01.reset_index()
        
        lista_colunas_s1 = planilha01.columns.tolist()

        cod1 = (lista_colunas_s1[3]) # codigo
        qtd1 = (lista_colunas_s1[1]) # quantidade
        x = (planilha01.loc[0:,[cod1]])
        # print(x)

        
        for index, row in planilha01.iterrows():
            codigo = (planilha01.loc[0:,[cod1]])
            quantidade = (planilha01.loc[0:,[qtd1]])
            list_qtd.append(quantidade)
            list_cod.append(codigo)
       
    def get_sec_sheet(self):
        get_version_new = QFileDialog.getOpenFileName(self, 'Open file', 'D:\Programação\CompareSheets\compare-sheets') # pega o arquivo
        self.filename2.setText(get_version_new[0])

        file_path2 = (get_version_new[0])
        paths.append(file_path2)

        updated_sheet = pd.read_excel(file_path2)
        sheet2.append(updated_sheet)
       
        planilha02 = pd.read_excel(file_path2)
        planilha02 = planilha02.reset_index()

        lista_colunas_s2 = planilha02.columns.to_list()

        cod2 = (lista_colunas_s2[3])
        qtd2 = (lista_colunas_s2[1])
        
       

        codigo2 = (planilha02.loc[0:,[cod2]])
        list_cod2.append(codigo2)
        # for index, row in planilha02.iterrows():
        #     quantidade2  = (planilha02.loc[0:,[qtd2]])
        # list_qtd2.append(quantidade2)
        # list_cod2.append(codigo2)

    def comparar(self):
        print(list_qtd[2],  " ------------------ ")
        # x = (planilha01.loc[0:,[cod1]])
        
        # cod_list = []
         
        # for row in (cod1):
        #     cod_list.append(row)
        # print(x)


        # for row in cod1:
        #     if row in cod2:
        #         if row not in n_cod:
        #             n_cod.append(row)
        # print(n_cod)


        # print(planilha01)
        # for i in planilha01:
        #     print(planilha01.loc[[i]])
        # for row in range(1, 8):
            #     for col in range(1, 6 ):
        #         char = get_column_letter(col)
               
#         finalList = []
#         # fileone = sheet1[0].readlines() --- # exemplo
#         lines = []
#         lines2 =[]
#         planilha01  = pd.DataFrame(sheet1[0])
#         planilha02 = pd.DataFrame(sheet2[0])

#         with open(tostring(paths[0]), 'r') as p1, open(tostring(paths[1] , 'r'))  as p2:
#             sheet_one = p1.readlines
#             sheet_two = p2.readlines

#             for line in sheet_one:
#                 if line not in sheet_two:
#                     finalList.append(line)
            
#             output = pd.ExcelWriter('output.xlsx',  engine='xlsxwriter')
#         # print(f"{paths[0].readlines()}  AQUI READDDDDDDDDDDDDD ")

#         print(type(paths[0]))
#         print(planilha01)
        
        # linhas planiilhas s
        # for line in planilha01:
        #     print(line)
        # print('acabou1')
            # lines.append(list(line.iloc[0,0]))
        # linhas planilhas 2
        # for line2 in planilha02:
        #     print(line2)
        # print('acabou2')

            # lines2.append(list(line2.iloc[0,0]))

        # for lineTotal in lines2:
        #     if lineTotal not in lines:
        #         writer = pd.ExcelWriter('output.xlsx', engine ='xlsxwriter')
        #         writer.save()
                # escrever a nova planilha com as diferenças
               

        # pegar linha por linha do arquivo
        



    #  equivalente a readline 

#      lines = []

#     In [387]: for line in pd.read_csv('./data_sample.tsv', encoding='utf-8', header=None, chunksize=1):
#     lines.append(line.iloc[0,0])


        
# iterando sobre a planilha selecionada



# #                                                         - OPEN PY XL VERSION

# class MainWindow(QDialog):
#     paths == []
#     def __init__(self):
#         super(MainWindow, self).__init__()
#         super().__init__()
        
#         loadUi("D:\Programação\CompareSheets\compare-sheets\\up_file-screen.ui", self)
#         self.browse.clicked.connect(self.get_first_sheet)
#         self.browse2.clicked.connect(self.get_sec_sheet)
#         self.comparar_btn.clicked.connect(self.comparar)

#     def get_first_sheet(self):
#         get_version_old = QFileDialog.getOpenFileName(self, 'Open file', 'D:\Programação\CompareSheets\compare-sheets')         # pega o arquivo
#         self.filename.setText(get_version_old[0])  
#         file_path = (get_version_old[0]) # caminho 
#         paths.append(file_path)
        

#     def get_sec_sheet(self):
#         get_version_new = QFileDialog.getOpenFileName(self, 'Open file', 'D:\Programação\CompareSheets\compare-sheets') # pega o arquivo
#         self.filename2.setText(get_version_new[0])
#         file_path2 = (get_version_new[0]) # caminho 
#         paths.append(file_path2)

#     def comparar(self):
#         wb1 = load_workbook(paths[0])
#         ws1 = wb1.active
       
#         wb2 = load_workbook(paths[1])
#         ws2 = wb2.active

#         for row in range(1, 8):
#             for col in range(1, 6 ):
#                 char = get_column_letter(col)
#                 print(ws1[char + str(row)].value)











app = QApplication(sys.argv)
mainwindow = MainWindow()
widget = QtWidgets.QStackedWidget()
widget.addWidget(mainwindow)
widget.setFixedWidth(800)
widget.setFixedHeight(800)
widget.show()
sys.exit(app.exec_())
